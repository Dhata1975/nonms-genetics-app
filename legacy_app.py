from __future__ import annotations
import io
import re
import zipfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional
from collections import defaultdict

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether
)

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"

# Version 8 modular dataset handling
# Drop any .txt panel into /data and the app will discover it automatically.
# Add a display name below only when you want a prettier label than the file stem.
PRIORITY_DATASET_ORDER = [
    "Immunometabolic Core.txt",
    "MPOA Network.txt",
    "Evolutionary Immune Network.txt",
    "ERAP2.txt",
    "DIO Thermoregulation.txt",
    "MS Attributes.txt",
    "Methylation.txt",
    "Homocysteine.txt",
    "Dysautonomia.txt",
    "Autonomic Loop.txt",
    "Molecular Mimicry.txt",
]

DISPLAY_NAMES = {
    "MS Attributes.txt": "MS GWAS",
    "SAM E Vulnerability.txt": "SAM-e Vulnerability",
    "Stress Event.txt": "Stress Nexus Event",
    "Small Vessel Disorder.txt": "Small Vessel Disorder",
    "Autonomic Loop.txt": "Autonomic Loop",
    "Molecular Mimicry.txt": "Molecular Mimicry",
    "CSVD.txt": "CSVD",
    "Homocysteine.txt": "Homocysteine",
    "Mold Fungus.txt": "Mold / Fungus",
    "Tinea Versicolor.txt": "Tinea Versicolor",
    "H Pylori.txt": "H. pylori",
    "Cardiomegaly.txt": "Cardiomegaly",
    "Inverted T-waves.txt": "T-Waves",
    "Low B12.txt": "Low B12",
    "Periodontal Disease.txt": "Periodontal Disease",
    "Methylation.txt": "Methylation",
    "Dysautonomia.txt": "Dysautonomia",
    "Immunometabolic Core.txt": "Immunometabolic Core",
    "MPOA Network.txt": "MPOA Network",
    "Evolutionary Immune Network.txt": "Evolutionary Immune Network",
    "ERAP2.txt": "ERAP2 / Ancient Selection",
    "DIO Thermoregulation.txt": "DIO Thermoregulation",
}



@dataclass
class Dataset:
    file_name: str
    category: str
    frame: pd.DataFrame


def normalize_category_name(name: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "_", name).strip("_")[:31] or "Sheet"


def make_unique_sheet_name(base: str, used_names: set[str]) -> str:
    base = (base or "Sheet")[:31]
    if base not in used_names:
        used_names.add(base)
        return base
    i = 2
    while True:
        suffix = f"_{i}"
        candidate = f"{base[:31-len(suffix)]}{suffix}"
        if candidate not in used_names:
            used_names.add(candidate)
            return candidate
        i += 1


def clean_excel_value(value):
    if value is None:
        return None
    if isinstance(value, (int, float, bool)):
        return value
    s = str(value)
    # Remove control characters that can break openpyxl
    s = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", s)
    return s


def parse_marker_token(token: str) -> tuple[Optional[str], Optional[str], str, str]:
    token = str(token).strip()
    if not token:
        return None, None, "blank", "Blank marker"

    if ";" in token:
        rsids = re.findall(r"rs\d+", token)
        if len(rsids) == 1:
            return rsids[0], None, "composite", "Composite marker; manual review needed"
        return None, None, "composite", "Composite marker / haplotype entry; manual review needed"

    if token.startswith("DRB") or token.startswith("A*") or "*" in token:
        return None, None, "hla", "HLA allele entry; not directly testable from standard Ancestry raw SNP export"

    m = re.match(r"^(rs\d+)-([A-Za-z\?]+)$", token)
    if m:
        rsid, allele = m.groups()
        allele = allele.upper()
        if allele == "?":
            return rsid, None, "rsid_unknown_allele", "Marker present but allele unspecified"
        if re.fullmatch(r"[ACGTID]+", allele) and len(allele) == 1:
            return rsid, allele, "rsid_single_allele", "Allele-aware comparison available"
        return rsid, allele, "rsid_multibase_allele", "Non-single-base allele; compare with caution"

    m = re.match(r"^(rs\d+)$", token)
    if m:
        return m.group(1), None, "rsid_only", "Presence / absence comparison available"

    if token.startswith("chr"):
        return None, None, "coordinate_marker", "Coordinate-based marker; not directly matched because raw file is rsID-based"

    return None, None, "other", "Unrecognized marker format"


def normalize_dataset_lines(path: Path) -> list[str]:
    raw_lines = path.read_text(encoding="utf-8", errors="ignore").splitlines()
    cleaned = []
    i = 0
    marker_hint = re.compile(r"(\s|\t)(rs\S+|DRB\S+|A\*\S+|chr\S+|kgp\S+)$")
    while i < len(raw_lines):
        line = raw_lines[i].strip()
        if not line:
            i += 1
            continue
        if marker_hint.search(line) or line.startswith("#") or "rsID/SNP" in line or "TRAIT" in line:
            cleaned.append(line)
            i += 1
            continue
        if i + 1 < len(raw_lines):
            nxt = raw_lines[i + 1].strip()
            if marker_hint.search(nxt):
                cleaned.append(f"{line} {nxt}")
                i += 2
                continue
        cleaned.append(line)
        i += 1
    return cleaned


def parse_generic_dataset(path: Path) -> Dataset:
    rows = []
    fixed_category = DISPLAY_NAMES.get(path.name, path.stem)
    for line in normalize_dataset_lines(path):
        line = line.strip()
        if not line:
            continue
        if line.startswith("#") or "rsID/SNP" in line or "TRAIT" in line:
            continue

        entry_id = None
        trait_label = fixed_category
        marker = None

        parts = re.split(r"\t+", line)

        if len(parts) >= 3 and parts[0].strip().isdigit():
            entry_id = int(parts[0].strip())
            trait_label = parts[1].strip() or fixed_category
            marker = parts[2].strip()
        else:
            m = re.match(r"^(\d+)\s+(.+?)\s+(rs\S+|DRB\S+|A\*\S+|chr\S+|kgp\S+)$", line)
            if m:
                entry_id = int(m.group(1))
                trait_label = m.group(2).strip() or fixed_category
                marker = m.group(3).strip()
            else:
                if len(parts) >= 2:
                    trait_label = parts[0].strip() or fixed_category
                    marker = parts[1].strip()
                else:
                    marker = line.strip()

        rsid, allele, marker_type, note = parse_marker_token(marker)
        rows.append({
            "entry_id": entry_id,
            "category": fixed_category,
            "trait_label": trait_label,
            "raw_marker": marker,
            "rsid": rsid,
            "listed_allele": allele,
            "marker_type": marker_type,
            "note": note,
        })

    df = pd.DataFrame(rows)
    if df.empty:
        df = pd.DataFrame(columns=["entry_id", "category", "trait_label", "raw_marker", "rsid", "listed_allele", "marker_type", "note"])
    return Dataset(path.name, fixed_category, df)


def discover_dataset_files() -> list[Path]:
    """Discover every .txt SNP panel in /data.

    Version 9 is modular: future pathway panels can be added by dropping
    a new .txt file into /data. No Python edit is required unless you want
    a custom display name or priority position.
    """
    if not DATA_DIR.exists():
        return []

    txt_files = {p.name: p for p in DATA_DIR.glob("*.txt") if p.is_file()}
    ordered = []
    for name in PRIORITY_DATASET_ORDER:
        if name in txt_files:
            ordered.append(txt_files.pop(name))

    ordered.extend(sorted(txt_files.values(), key=lambda p: DISPLAY_NAMES.get(p.name, p.stem).lower()))
    return ordered


def load_all_datasets() -> list[Dataset]:
    return [parse_generic_dataset(path) for path in discover_dataset_files()]


def split_genotype(genotype: str) -> tuple[str, str]:
    genotype = (genotype or "").strip().upper()
    if genotype in {"", "--", "0", "00"}:
        return "", ""
    if len(genotype) == 1:
        return genotype, ""
    if len(genotype) == 2:
        return genotype[0], genotype[1]
    return "", ""


def parse_dna_file(uploaded_file) -> pd.DataFrame:
    """Parse AncestryDNA, 23andMe, and MyHeritage raw DNA files.

    Supported input patterns:
    - AncestryDNA: rsid, chromosome, position, allele1, allele2
    - 23andMe: rsid, chromosome, position, genotype
    - MyHeritage MHv1.0: RSID,CHROMOSOME,POSITION,RESULT
    """
    rows = []
    content = uploaded_file.getvalue().decode("utf-8-sig", errors="ignore").splitlines()

    detected_format = None
    for header_line in content[:50]:
        lower = header_line.lower()
        if "fileformat=myheritage" in lower or "format=mhv" in lower:
            detected_format = "MyHeritage"
            break
        if "23andme" in lower:
            detected_format = "23andMe"
            break
        if lower.startswith("rsid\tchromosome\tposition\tallele1\tallele2"):
            detected_format = "AncestryDNA"
            break

    for raw_line in content:
        line = raw_line.strip()
        if not line:
            continue

        lower = line.lower()
        if line.startswith("#"):
            continue

        # MyHeritage CSV header and rows: RSID,CHROMOSOME,POSITION,RESULT
        if lower.startswith("rsid,chromosome,position,result"):
            detected_format = "MyHeritage"
            continue

        if detected_format == "MyHeritage" or "," in line:
            parts_csv = [p.strip().strip('"') for p in line.split(",")]
            if len(parts_csv) == 4:
                rsid, chrom, pos, genotype = parts_csv
                if not (rsid.startswith("rs") or rsid.startswith("i")):
                    continue
                try:
                    pos = int(pos)
                except ValueError:
                    continue
                a1, a2 = split_genotype(genotype)
                detected_format = detected_format or "MyHeritage"
                rows.append((rsid.strip(), chrom.strip(), pos, a1, a2))
                continue

        parts = line.split("\t")
        if len(parts) == 1:
            parts = re.split(r"\s+", line)

        # 23andMe: rsid chrom position genotype
        if len(parts) == 4:
            rsid, chrom, pos, genotype = parts
            if not (rsid.startswith("rs") or rsid.startswith("i")):
                continue
            try:
                pos = int(pos)
            except ValueError:
                continue
            a1, a2 = split_genotype(genotype)
            detected_format = detected_format or "23andMe"
            rows.append((rsid.strip(), chrom.strip(), pos, a1, a2))

        # AncestryDNA: rsid chrom position allele1 allele2
        elif len(parts) == 5:
            rsid, chrom, pos, a1, a2 = parts
            if not (rsid.startswith("rs") or rsid.startswith("i")):
                continue
            try:
                pos = int(pos)
            except ValueError:
                continue
            a1 = a1.strip().upper()
            a2 = a2.strip().upper()
            if a1 == "0":
                a1 = ""
            if a2 == "0":
                a2 = ""
            detected_format = detected_format or "AncestryDNA"
            rows.append((rsid.strip(), chrom.strip(), pos, a1, a2))

    df = pd.DataFrame(rows, columns=["rsid", "chromosome", "position", "allele1", "allele2"])
    if df.empty:
        return pd.DataFrame(columns=["rsid", "chromosome", "position", "allele1", "allele2", "genotype", "source"])

    df["genotype"] = df["allele1"].fillna("") + df["allele2"].fillna("")
    df["source"] = detected_format or "Unknown"
    return df

def compare_dataset(dataset: Dataset, ancestry_df: pd.DataFrame) -> pd.DataFrame:
    df = dataset.frame.copy()
    if ancestry_df.empty:
        return df.assign(
            in_ancestry=False, chromosome=None, position=None, allele1=None, allele2=None,
            genotype=None, listed_allele_copies=None, zygosity=None,
            comparison_status="No DNA data loaded", manual_review="Yes"
        )

    anc = ancestry_df.drop_duplicates("rsid").set_index("rsid")
    result_rows = []

    for _, row in df.iterrows():
        rsid = row["rsid"]
        allele = row["listed_allele"]
        marker_type = row["marker_type"]

        if not rsid:
            result_rows.append({
                **row.to_dict(),
                "in_ancestry": False,
                "chromosome": None,
                "position": None,
                "allele1": None,
                "allele2": None,
                "genotype": None,
                "listed_allele_copies": None,
                "zygosity": None,
                "comparison_status": "Not directly comparable",
                "manual_review": "Yes",
            })
            continue

        if rsid not in anc.index:
            result_rows.append({
                **row.to_dict(),
                "in_ancestry": False,
                "chromosome": None,
                "position": None,
                "allele1": None,
                "allele2": None,
                "genotype": None,
                "listed_allele_copies": None,
                "zygosity": None,
                "comparison_status": "rsID not present in uploaded DNA file",
                "manual_review": "Maybe",
            })
            continue

        rec = anc.loc[rsid]
        genotype = f"{rec['allele1']}{rec['allele2']}"

        status = ""
        copies = None
        zygosity = None
        manual_review = "No"

        if marker_type == "rsid_single_allele":
            copies = int(rec["allele1"] == allele) + int(rec["allele2"] == allele)
            if copies == 2:
                status = "Listed allele present"
                zygosity = "Homozygous listed allele"
            elif copies == 1:
                status = "Listed allele present"
                zygosity = "Heterozygous"
            else:
                status = "Listed allele absent"
                zygosity = "Listed allele absent"
        elif marker_type == "rsid_only":
            status = "Marker present"
            zygosity = "Genotype observed"
        elif marker_type == "rsid_unknown_allele":
            status = "Present, but listed allele unspecified"
            manual_review = "Yes"
        elif marker_type == "rsid_multibase_allele":
            copies = int(rec["allele1"] == allele) + int(rec["allele2"] == allele)
            status = "Possible exact multibase match" if copies > 0 else "No exact multibase match"
            zygosity = "Contains multibase allele" if copies > 0 else "Multibase allele not detected"
            manual_review = "Yes"
        else:
            status = "Present, manual review needed"
            manual_review = "Yes"

        result_rows.append({
            **row.to_dict(),
            "in_ancestry": True,
            "chromosome": rec["chromosome"],
            "position": int(rec["position"]),
            "allele1": rec["allele1"],
            "allele2": rec["allele2"],
            "genotype": genotype,
            "listed_allele_copies": copies,
            "zygosity": zygosity,
            "comparison_status": status,
            "manual_review": manual_review,
        })

    return pd.DataFrame(result_rows)


def build_summary(all_results: pd.DataFrame) -> pd.DataFrame:
    summaries = []
    for category, grp in all_results.groupby("category", sort=True):
        total = len(grp)
        directly_comparable = int(grp["marker_type"].isin(["rsid_single_allele", "rsid_only"]).sum())
        found = int(grp["in_ancestry"].fillna(False).sum())
        allele_present = int((grp["comparison_status"] == "Listed allele present").sum())
        marker_present = int((grp["comparison_status"] == "Marker present").sum())
        absent = int((grp["comparison_status"] == "Listed allele absent").sum())
        missing = int((grp["comparison_status"] == "rsID not present in uploaded DNA file").sum())
        manual = int((grp["manual_review"] != "No").sum())
        comparable_found = int(((grp["marker_type"].isin(["rsid_single_allele", "rsid_only"])) & (grp["in_ancestry"].fillna(False))).sum())
        coverage_pct = round(100 * comparable_found / directly_comparable, 1) if directly_comparable else None
        hit_pct = round(100 * (allele_present + marker_present) / comparable_found, 1) if comparable_found else None

        summaries.append({
            "category": category,
            "rows": total,
            "directly_comparable_rows": directly_comparable,
            "present_in_ancestry": found,
            "listed_allele_present": allele_present,
            "marker_present_no_allele": marker_present,
            "listed_allele_absent": absent,
            "missing_from_ancestry": missing,
            "manual_review_rows": manual,
            "coverage_pct_of_comparable": coverage_pct,
            "match_pct_when_present": hit_pct,
        })

    return pd.DataFrame(summaries).sort_values(["category"]).reset_index(drop=True)


def build_overall_metrics(summary_df: pd.DataFrame, all_results: pd.DataFrame) -> dict:
    if summary_df.empty:
        return {}
    return {
        "Datasets loaded": int(summary_df["category"].nunique()),
        "Total marker rows": int(len(all_results)),
        "Directly comparable rows": int(summary_df["directly_comparable_rows"].sum()),
        "Present in volunteer file": int(summary_df["present_in_ancestry"].sum()),
        "Allele / marker hits": int((summary_df["listed_allele_present"] + summary_df["marker_present_no_allele"]).sum()),
        "Manual review rows": int(summary_df["manual_review_rows"].sum()),
    }


def make_excel_report(summary_df: pd.DataFrame, all_results: pd.DataFrame, ancestry_df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "README"
    used_names = {"README"}

    navy = PatternFill("solid", fgColor="13263A")
    white_bold = Font(color="FFFFFF", bold=True)

    ws0.merge_cells("A1:F1")
    ws0["A1"] = "NONMS Genetics Comparison Report"
    ws0["A1"].fill = navy
    ws0["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    ws0["A3"] = "Generated"
    ws0["B3"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    ws0["A5"] = "Important"
    ws0["B5"] = "This workbook is a pattern-comparison tool. It is not a diagnosis, clinical interpretation, or validated polygenic risk score."
    ws0["A7"] = "Input file rows"
    ws0["B7"] = len(ancestry_df)
    ws0["A8"] = "Marker rows reviewed"
    ws0["B8"] = len(all_results)
    ws0.column_dimensions["A"].width = 22
    ws0.column_dimensions["B"].width = 100
    ws0.sheet_view.showGridLines = False

    def add_df_sheet(name: str, df: pd.DataFrame):
        safe_name = make_unique_sheet_name(name[:31], used_names)
        ws = wb.create_sheet(safe_name)
        if df.empty:
            ws["A1"] = "No rows"
            return ws
        for c_idx, col in enumerate(df.columns, 1):
            cell = ws.cell(1, c_idx, clean_excel_value(col))
            cell.fill = navy
            cell.font = white_bold
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for r_idx, row in enumerate(df.itertuples(index=False), 2):
            for c_idx, val in enumerate(row, 1):
                ws.cell(r_idx, c_idx, clean_excel_value(val))
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(df.shape[1])}{ws.max_row}"
        ws.sheet_view.showGridLines = False
        for idx, col in enumerate(df.columns, start=1):
            width = max(len(str(col)) + 2, 12)
            sample_vals = [clean_excel_value(v) for v in df[col].head(200).tolist()]
            sample_vals = ["" if v is None else str(v) for v in sample_vals]
            if sample_vals:
                width = max(width, min(42, int(max(len(v) for v in sample_vals) * 0.95) + 2))
            ws.column_dimensions[get_column_letter(idx)].width = min(max(width, 12), 45)
        return ws

    add_df_sheet("Category_Summary", summary_df)
    add_df_sheet("All_Results", all_results)
    add_df_sheet("Matched_Only", all_results[all_results["comparison_status"].isin(["Listed allele present", "Marker present", "Possible exact multibase match"])])
    add_df_sheet("Manual_Review", all_results[all_results["manual_review"] != "No"])

    for category in summary_df["category"].tolist():
        cat_df = all_results[all_results["category"] == category].copy()
        add_df_sheet(normalize_category_name(category), cat_df)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()




def draw_pdf_header_dark(canvas, doc):
    canvas.saveState()
    canvas.setFillColor(colors.HexColor("#0A0A0A"))
    canvas.rect(0, 0, doc.pagesize[0], doc.pagesize[1], fill=1, stroke=0)
    canvas.setFillColor(colors.HexColor("#D4AF37"))
    canvas.rect(0.5 * inch, doc.pagesize[1] - 0.8 * inch, doc.pagesize[0] - inch, 0.08 * inch, fill=1, stroke=0)
    canvas.setFillColor(colors.white)
    canvas.setFont("Helvetica-Bold", 9)
    canvas.drawString(0.6 * inch, doc.pagesize[1] - 0.55 * inch, "AREA 76 // NONMS COMMAND CENTER")
    canvas.drawRightString(doc.pagesize[0] - 0.6 * inch, 0.45 * inch, f"Page {doc.page}")
    canvas.restoreState()


def draw_pdf_header_light(canvas, doc):
    canvas.saveState()
    canvas.setFillColor(colors.white)
    canvas.rect(0, 0, doc.pagesize[0], doc.pagesize[1], fill=1, stroke=0)
    canvas.setStrokeColor(colors.HexColor("#D4AF37"))
    canvas.setLineWidth(1)
    canvas.line(0.65 * inch, doc.pagesize[1] - 0.62 * inch, doc.pagesize[0] - 0.65 * inch, doc.pagesize[1] - 0.62 * inch)
    canvas.setFillColor(colors.HexColor("#222222"))
    canvas.setFont("Helvetica-Bold", 9)
    canvas.drawString(0.7 * inch, doc.pagesize[1] - 0.48 * inch, "NONMS Genetics Report")
    canvas.drawRightString(doc.pagesize[0] - 0.7 * inch, 0.48 * inch, f"Page {doc.page}")
    canvas.restoreState()





# ----------------------------
# Version 9 volunteer education layer
# ----------------------------

PATHWAY_PLAIN_ENGLISH = {
    "Immunometabolic Core": {
        "plain": "This section looks at genes connected to how the body links immune activity with energy use.",
        "why": "NONMS studies this because inflammation, fatigue, repair, and metabolism often overlap in chronic illness."
    },
    "MPOA Network": {
        "plain": "This section looks at genes connected to temperature control, stress response, sleep, and automatic body functions.",
        "why": "NONMS studies this because the brain's control systems may influence heat sensitivity, fatigue, and recovery."
    },
    "Evolutionary Immune Network": {
        "plain": "This section looks at genes that may connect ancient immune survival pressures with modern immune behavior.",
        "why": "NONMS studies this because some immune traits that helped humans survive infections may also shape autoimmune risk today."
    },
    "ERAP2 / Ancient Selection": {
        "plain": "This section looks at ERAP2, a gene involved in how the immune system shows threat signals to immune cells.",
        "why": "ERAP2 is included because it has been discussed in research about past infection pressure and immune tradeoffs."
    },
    "DIO Thermoregulation": {
        "plain": "This section looks at genes involved in thyroid signaling and tissue-level temperature regulation.",
        "why": "NONMS studies this because heat sensitivity and energy regulation are major questions in MS and related conditions."
    },
    "MS GWAS": {
        "plain": "This section compares your file with published MS-associated research markers.",
        "why": "These markers do not diagnose MS. They help researchers look for patterns across many people."
    },
    "Methylation": {
        "plain": "This section looks at genes involved in B vitamins, folate, methylation, and repair chemistry.",
        "why": "NONMS studies this because methylation affects homocysteine, DNA regulation, nerve health, and cellular repair."
    },
    "Homocysteine": {
        "plain": "This section looks at markers connected to homocysteine handling.",
        "why": "Homocysteine matters because it can affect blood vessels, inflammation, and nervous system stress."
    },
    "Dysautonomia": {
        "plain": "This section looks at genes connected to automatic body functions.",
        "why": "These pathways may matter for heart rate, blood pressure, temperature control, digestion, and fatigue."
    },
    "Autonomic Loop": {
        "plain": "This section looks at genes connected to stress chemistry and automatic nervous system signaling.",
        "why": "NONMS studies this because stress signals can change energy use, immune activity, and symptom intensity."
    },
    "Molecular Mimicry": {
        "plain": "This section looks at immune markers related to the idea that infections may sometimes confuse immune recognition.",
        "why": "This is a research concept, not a conclusion about any one volunteer."
    },
}

def safe_num(value, default=0):
    try:
        if pd.isna(value):
            return default
        return float(value)
    except Exception:
        return default

def get_top_categories_for_summary(summary_df: pd.DataFrame, max_items: int = 5) -> list[str]:
    if summary_df is None or summary_df.empty:
        return []
    ranked = summary_df.copy()
    ranked["hits_total"] = ranked["listed_allele_present"].fillna(0) + ranked["marker_present_no_allele"].fillna(0)
    preferred = [
        "Immunometabolic Core",
        "MPOA Network",
        "Evolutionary Immune Network",
        "DIO Thermoregulation",
        "ERAP2 / Ancient Selection",
        "MS GWAS",
        "Methylation",
        "Dysautonomia",
    ]
    ordered = []
    for cat in preferred:
        if cat in ranked["category"].values:
            ordered.append(cat)
    remaining = ranked.sort_values(["hits_total", "present_in_ancestry"], ascending=False)["category"].tolist()
    for cat in remaining:
        if cat not in ordered:
            ordered.append(cat)
    return ordered[:max_items]

def build_plain_english_summary_paragraphs(summary_df: pd.DataFrame, all_results: pd.DataFrame) -> list[str]:
    total_datasets = int(summary_df["category"].nunique()) if summary_df is not None and not summary_df.empty else 0
    total_markers = int(len(all_results)) if all_results is not None else 0
    total_present = int(summary_df["present_in_ancestry"].sum()) if summary_df is not None and not summary_df.empty else 0

    paragraphs = [
        "Thank you for participating in the NONMS research project. Your DNA is one piece of a much larger puzzle. This report looks for research markers in pathways related to immunity, energy regulation, temperature control, methylation, and nervous system balance.",
        "This report is not a diagnosis. It cannot tell you whether you have MS or any other disease. A DNA marker is a clue, not a verdict. The purpose of this report is to help researchers compare patterns across volunteers.",
        f"In this report, the app reviewed {total_datasets} research panels containing {total_markers} marker rows. Your uploaded DNA file contained data for {total_present} of those marker rows. When a marker is missing, it often means the DNA company did not test that exact SNP.",
        "The most important idea is simple: one gene does not tell the whole story. NONMS looks at groups of genes working together in biological pathways. Those pathways may help researchers ask better questions about inflammation, fatigue, heat sensitivity, repair, and recovery.",
    ]
    return paragraphs

def build_research_moment_text() -> list[str]:
    return [
        "Research Moment: DNA is not destiny. Your genes are more like a parts list. Health is influenced by genes, environment, infections, stress, nutrition, sleep, and many other factors.",
        "Research Moment: A SNP is a single-letter difference in DNA. Most SNPs are common and harmless. Researchers study them because patterns across many SNPs may help explain how biological systems behave.",
        "Research Moment: A missing result usually does not mean the marker is absent from your body. It usually means that specific DNA testing company did not include that SNP on its chip.",
        "Research Moment: NONMS is especially interested in pathways, not isolated genes. Pathways help show how energy, immunity, temperature control, and repair may interact."
    ]

def add_volunteer_summary_pages(story, styles, summary_df: pd.DataFrame, all_results: pd.DataFrame, dark_mode: bool):
    """Add Version 9 plain-English pages before the technical report."""
    body_style = styles["BodyMode"]
    heading_style = styles["HeadingMode"]
    small_style = styles["SmallMode"]

    story.append(Paragraph("Welcome to the NONMS Project", heading_style))
    for para in build_plain_english_summary_paragraphs(summary_df, all_results):
        story.append(Paragraph(para, body_style))
        story.append(Spacer(1, 0.08 * inch))

    story.append(Paragraph("How to Read This Report", heading_style))
    how_to = [
        "<b>Found / present</b> means the marker was seen in the uploaded DNA file.",
        "<b>Missing</b> usually means the DNA company did not test that exact SNP.",
        "<b>High alignment</b> means many markers in that panel were found in the uploaded file. It does not mean high disease risk.",
        "<b>Manual review</b> means the marker may need closer inspection because of HLA markers, composite entries, unknown alleles, or chip limitations.",
        "<b>Coverage</b> tells us how much of a panel could actually be checked from the uploaded DNA file."
    ]
    for item in how_to:
        story.append(Paragraph("• " + item, body_style))
    story.append(Spacer(1, 0.12 * inch))

    moments = build_research_moment_text()
    for moment in moments[:2]:
        story.append(Paragraph(moment, small_style))
        story.append(Spacer(1, 0.06 * inch))

    story.append(PageBreak())

    story.append(Paragraph("Your Genetic Story", heading_style))
    story.append(Paragraph(
        "The sections below explain the main research pathways in plain English. These are educational summaries. The full technical tables begin after this introduction.",
        body_style
    ))
    story.append(Spacer(1, 0.10 * inch))

    top_categories = get_top_categories_for_summary(summary_df, max_items=8)
    for cat in top_categories:
        info = PATHWAY_PLAIN_ENGLISH.get(cat, {
            "plain": "This section looks at one of the research pathways included in the NONMS genetics engine.",
            "why": "It is included to help researchers compare patterns across volunteers."
        })
        if summary_df is not None and not summary_df.empty and cat in summary_df["category"].values:
            row = summary_df[summary_df["category"] == cat].iloc[0]
            present = int(row["present_in_ancestry"])
            rows = int(row["rows"])
            hits = int(row["listed_allele_present"] + row["marker_present_no_allele"])
            coverage = "-" if pd.isna(row["coverage_pct_of_comparable"]) else f"{row['coverage_pct_of_comparable']:.1f}%"
            signal = str(row.get("signal_call", ""))
        else:
            present = rows = hits = 0
            coverage = "-"
            signal = ""

        story.append(KeepTogether([
            Paragraph(cat, heading_style),
            Paragraph(info["plain"], body_style),
            Paragraph(info["why"], body_style),
            Paragraph(f"Research snapshot: {present} of {rows} marker rows were present in the uploaded file; {hits} positive marker signal(s) were found. Coverage: {coverage}. Signal call: {signal}.", small_style),
            Spacer(1, 0.10 * inch),
        ]))

    story.append(PageBreak())

    story.append(Paragraph("What We Studied", heading_style))
    studied = [
        ("Immune signaling", "How the body recognizes threats and communicates danger."),
        ("Energy regulation", "How cells decide whether to spend energy, conserve energy, or repair."),
        ("Temperature control", "How the body manages heat, thyroid signaling, and stress physiology."),
        ("Methylation and B vitamins", "How the body handles repair chemistry, homocysteine, and cellular regulation."),
        ("Nervous system balance", "How automatic body functions such as heart rate, temperature, sleep, and stress responses may be coordinated."),
        ("Evolutionary immune traits", "How ancient survival pressures may have shaped immune patterns that researchers still study today."),
    ]
    for title, desc in studied:
        story.append(Paragraph(f"<b>{title}</b>: {desc}", body_style))
        story.append(Spacer(1, 0.06 * inch))

    for moment in moments[2:]:
        story.append(Spacer(1, 0.06 * inch))
        story.append(Paragraph(moment, small_style))

    story.append(Spacer(1, 0.12 * inch))
    story.append(Paragraph(
        "The next section begins the technical report. It keeps the detailed data so volunteers, researchers, and clinicians can go deeper after reading the summary.",
        body_style
    ))
    story.append(PageBreak())


def make_pdf_report(summary_df: pd.DataFrame, all_results: pd.DataFrame, volunteer_filename: str, report_style: str = "Command Center (dark)") -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=0.7 * inch,
        rightMargin=0.7 * inch,
        topMargin=0.9 * inch,
        bottomMargin=0.7 * inch,
    )

    dark_mode = report_style == "Command Center (dark)"
    styles = getSampleStyleSheet()

    if dark_mode:
        title_color = colors.HexColor("#D4AF37")
        body_color = colors.white
        head_fill = colors.HexColor("#D4AF37")
        head_text = colors.black
        row_fill = colors.HexColor("#111111")
        row_text = colors.white
        border_color = colors.HexColor("#555555")
        note_label = "AREA 76 // NONMS COMMAND CENTER"
        title_text = "GENETICS SIGNAL REPORT"
        page_fn = draw_pdf_header_dark
        cover_subtitle = "Volunteer upload processed through bundled MS and pathway marker sets. This report is structured as a command summary: high-level signal first, drill-down second."
    else:
        title_color = colors.HexColor("#1F1F1F")
        body_color = colors.HexColor("#222222")
        head_fill = colors.HexColor("#EAEAEA")
        head_text = colors.black
        row_fill = colors.white
        row_text = colors.black
        border_color = colors.HexColor("#C8C8C8")
        note_label = "NONMS PRINT-FRIENDLY REPORT"
        title_text = "Genetics Pattern Comparison Report"
        page_fn = draw_pdf_header_light
        cover_subtitle = "Print-optimized volunteer comparison summary across bundled MS and pathway marker sets. Designed for paper readability and clinical-style review."

    styles.add(ParagraphStyle(name="TitleMode", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=20, leading=24, textColor=title_color, spaceAfter=8))
    styles.add(ParagraphStyle(name="BodyMode", parent=styles["BodyText"], fontName="Helvetica", fontSize=9.5, leading=13, textColor=body_color))
    styles.add(ParagraphStyle(name="HeadingMode", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=13, leading=16, textColor=title_color, spaceAfter=8, spaceBefore=8))
    styles.add(ParagraphStyle(name="SmallMode", parent=styles["BodyText"], fontName="Helvetica", fontSize=8.3, leading=11, textColor=body_color))

    story = []
    story.append(Paragraph(note_label, styles["SmallMode"]))
    story.append(Paragraph(title_text, styles["TitleMode"]))
    story.append(Paragraph(cover_subtitle, styles["BodyMode"]))
    story.append(Spacer(1, 0.18 * inch))

    meta_data = [
        ["Volunteer File", volunteer_filename],
        ["Generated (UTC)", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")],
        ["Report Mode", "Pattern comparison"],
        ["Style", "Area 76 dark" if dark_mode else "Print-friendly light"],
    ]
    meta_table = Table(meta_data, colWidths=[1.45 * inch, 4.85 * inch])
    meta_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), head_fill),
        ("TEXTCOLOR", (0, 0), (0, -1), head_text),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("BACKGROUND", (1, 0), (1, -1), row_fill),
        ("TEXTCOLOR", (1, 0), (1, -1), row_text),
        ("GRID", (0, 0), (-1, -1), 0.35, border_color),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 0.15 * inch))

    total_rows = int(len(all_results))
    total_datasets = int(summary_df["category"].nunique()) if not summary_df.empty else 0
    total_present = int(summary_df["present_in_ancestry"].sum()) if not summary_df.empty else 0
    total_hits = int((summary_df["listed_allele_present"] + summary_df["marker_present_no_allele"]).sum()) if not summary_df.empty else 0

    kpi_table = Table([
        ["Datasets", "Marker rows", "Present in file", "Positive hits"],
        [str(total_datasets), str(total_rows), str(total_present), str(total_hits)]
    ], colWidths=[1.45 * inch, 1.45 * inch, 1.45 * inch, 1.45 * inch])
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), head_fill),
        ("TEXTCOLOR", (0, 0), (-1, 0), head_text),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 1), (-1, 1), row_fill),
        ("TEXTCOLOR", (0, 1), (-1, 1), row_text),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.35, border_color),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 0.16 * inch))
    story.append(Paragraph("Mission Guardrails", styles["HeadingMode"]))
    story.append(Paragraph("Treat every match here as a pattern signal, not a conclusion. Marker presence can support hypothesis generation and cross-volunteer comparison, but it does not prove disease, rule disease out, or replace clinical interpretation.", styles["BodyMode"]))
    story.append(Spacer(1, 0.12 * inch))

    # V9: add plain-English volunteer education pages before the technical tables.
    add_volunteer_summary_pages(story, styles, summary_df, all_results, dark_mode)

    story.append(Paragraph("Executive Signal Board", styles["HeadingMode"]))
    summary_table_data = [["Category", "Rows", "Present", "Hits", "Coverage %", "Match %", "Signal Call"]]
    for _, row in summary_df.iterrows():
        summary_table_data.append([
            str(row["category"]),
            int(row["rows"]),
            int(row["present_in_ancestry"]),
            int(row["listed_allele_present"] + row["marker_present_no_allele"]),
            "-" if pd.isna(row["coverage_pct_of_comparable"]) else f'{row["coverage_pct_of_comparable"]:.1f}',
            "-" if pd.isna(row["match_pct_when_present"]) else f'{row["match_pct_when_present"]:.1f}',
            str(row.get("signal_call", "")),
        ])
    summary_col_widths = [2.1 * inch, 0.55 * inch, 0.7 * inch, 0.55 * inch, 0.75 * inch, 0.7 * inch, 1.0 * inch]
    table = Table(summary_table_data, repeatRows=1, colWidths=summary_col_widths)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), head_fill),
        ("TEXTCOLOR", (0, 0), (-1, 0), head_text),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 1), (-1, -1), row_fill),
        ("TEXTCOLOR", (0, 1), (-1, -1), row_text),
        ("GRID", (0, 0), (-1, -1), 0.25, border_color),
        ("FONTSIZE", (0, 0), (-1, -1), 8.4),
        ("LEADING", (0, 0), (-1, -1), 10.5),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(table)

    top_calls = summary_df.sort_values(["listed_allele_present", "marker_present_no_allele"], ascending=False).head(6)
    if not top_calls.empty:
        readout = " • ".join(
            f"{r['category']}: {int(r['listed_allele_present'] + r['marker_present_no_allele'])} positive row(s)"
            for _, r in top_calls.iterrows()
        )
        story.append(Spacer(1, 0.12 * inch))
        story.append(Paragraph("Command Readout" if dark_mode else "Top Positive Signal Readout", styles["HeadingMode"]))
        story.append(Paragraph(readout, styles["BodyMode"]))

    story.append(PageBreak())

    for category in summary_df["category"].tolist():
        grp_all = all_results[all_results["category"] == category].copy()
        cat_summary = summary_df[summary_df["category"] == category].iloc[0]

        story.append(Paragraph(category, styles["HeadingMode"]))
        story.append(Paragraph(
            f"Signal call: {cat_summary.get('signal_call', '')} | Coverage: "
            + ("-" if pd.isna(cat_summary["coverage_pct_of_comparable"]) else f"{cat_summary['coverage_pct_of_comparable']:.1f}%")
            + " | Match when present: "
            + ("-" if pd.isna(cat_summary["match_pct_when_present"]) else f"{cat_summary['match_pct_when_present']:.1f}%"),
            styles["BodyMode"]
        ))

        mini = Table([
            ["Rows", "Comparable", "Present", "Hits", "Missing", "Manual Review"],
            [
                int(cat_summary["rows"]),
                int(cat_summary["directly_comparable_rows"]),
                int(cat_summary["present_in_ancestry"]),
                int(cat_summary["listed_allele_present"] + cat_summary["marker_present_no_allele"]),
                int(cat_summary["missing_from_ancestry"]),
                int(cat_summary["manual_review_rows"]),
            ]
        ], colWidths=[0.85 * inch, 0.95 * inch, 0.8 * inch, 0.7 * inch, 0.8 * inch, 1.0 * inch])
        mini.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), head_fill),
            ("TEXTCOLOR", (0, 0), (-1, 0), head_text),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BACKGROUND", (0, 1), (-1, 1), row_fill),
            ("TEXTCOLOR", (0, 1), (-1, 1), row_text),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.25, border_color),
            ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(Spacer(1, 0.08 * inch))
        story.append(mini)
        story.append(Spacer(1, 0.10 * inch))

        pos = grp_all[grp_all["comparison_status"].isin(["Listed allele present", "Marker present", "Possible exact multibase match"])][
            ["raw_marker", "rsid", "listed_allele", "genotype", "comparison_status"]
        ].head(12)

        story.append(Paragraph("Positive Signal Rows", styles["HeadingMode"]))
        if pos.empty:
            story.append(Paragraph("No positive-signal rows surfaced in the top slice for this category.", styles["BodyMode"]))
        else:
            pos_data = [["Marker", "rsID", "Allele", "Genotype", "Status"]]
            for _, r in pos.iterrows():
                pos_data.append([
                    str(r["raw_marker"])[:28],
                    str(r["rsid"] or ""),
                    str(r["listed_allele"] or ""),
                    str(r["genotype"] or ""),
                    str(r["comparison_status"] or "")[:34],
                ])
            pos_table = Table(pos_data, repeatRows=1, colWidths=[1.8 * inch, 1.1 * inch, 0.55 * inch, 0.8 * inch, 2.0 * inch])
            pos_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), head_fill),
                ("TEXTCOLOR", (0, 0), (-1, 0), head_text),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("BACKGROUND", (0, 1), (-1, -1), row_fill),
                ("TEXTCOLOR", (0, 1), (-1, -1), row_text),
                ("GRID", (0, 0), (-1, -1), 0.25, border_color),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("LEADING", (0, 0), (-1, -1), 10),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            story.append(pos_table)

        story.append(Spacer(1, 0.10 * inch))
        story.append(Paragraph("Manual Review Flags", styles["HeadingMode"]))
        manual = grp_all[grp_all["manual_review"] != "No"][["raw_marker", "rsid", "marker_type", "note"]].head(10)
        if manual.empty:
            story.append(Paragraph("No manual-review rows surfaced in the top slice for this category.", styles["BodyMode"]))
        else:
            manual_data = [["Marker", "rsID", "Issue", "Note"]]
            for _, r in manual.iterrows():
                manual_data.append([
                    str(r["raw_marker"])[:24],
                    str(r["rsid"] or ""),
                    str(r["marker_type"] or ""),
                    str(r["note"] or "")[:48],
                ])
            man_table = Table(manual_data, repeatRows=1, colWidths=[1.7 * inch, 1.05 * inch, 1.0 * inch, 2.25 * inch])
            man_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), head_fill),
                ("TEXTCOLOR", (0, 0), (-1, 0), head_text),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("BACKGROUND", (0, 1), (-1, -1), row_fill),
                ("TEXTCOLOR", (0, 1), (-1, -1), row_text),
                ("GRID", (0, 0), (-1, -1), 0.25, border_color),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("LEADING", (0, 0), (-1, -1), 10),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            story.append(man_table)

        story.append(Spacer(1, 0.10 * inch))
        analyst_note = "Use the Excel export for full row-level detail, especially when a category contains HLA entries, composite markers, or unknown alleles."
        story.append(Paragraph("Analyst note: " + analyst_note, styles["SmallMode"]))

        if category != summary_df["category"].tolist()[-1]:
            story.append(PageBreak())

    doc.build(story, onFirstPage=page_fn, onLaterPages=page_fn)
    return buffer.getvalue()


def make_csv_zip(summary_df: pd.DataFrame, all_results: pd.DataFrame) -> bytes:
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("category_summary.csv", summary_df.to_csv(index=False))
        zf.writestr("all_results.csv", all_results.to_csv(index=False))
        for category in summary_df["category"].tolist():
            cat_df = all_results[all_results["category"] == category]
            zf.writestr(f"{normalize_category_name(category)}.csv", cat_df.to_csv(index=False))
    return buffer.getvalue()


@st.cache_resource
def load_datasets_cached():
    return load_all_datasets()


def inject_css():
    st.markdown("""
    <style>
    .stApp {
        background:
            radial-gradient(circle at top left, rgba(212, 175, 55, 0.09), transparent 28%),
            radial-gradient(circle at top right, rgba(62, 108, 168, 0.10), transparent 22%),
            linear-gradient(180deg, #06080D 0%, #0B0F17 45%, #090B10 100%);
        color: #F4F1E8;
    }
    .block-container {
        padding-top: 1.4rem;
        padding-bottom: 2rem;
        max-width: 1380px;
    }
    .hero-shell {
        border: 1px solid rgba(212, 175, 55, 0.20);
        background: linear-gradient(180deg, rgba(15,22,35,0.92), rgba(10,14,22,0.96));
        border-radius: 22px;
        padding: 1.4rem 1.5rem 1.15rem 1.5rem;
        box-shadow: 0 0 0 1px rgba(255,255,255,0.03) inset, 0 18px 44px rgba(0,0,0,0.35);
        margin-bottom: 1rem;
    }
    .hero-kicker {
        color: #D4AF37;
        font-size: 0.78rem;
        letter-spacing: 0.18em;
        text-transform: uppercase;
        font-weight: 700;
        margin-bottom: 0.35rem;
    }
    .hero-title {
        font-size: 2.55rem;
        line-height: 1.05;
        font-weight: 800;
        color: #F7F2E7;
        margin: 0 0 0.35rem 0;
    }
    .hero-sub {
        color: #B7C0D0;
        font-size: 1rem;
        margin-bottom: 1rem;
        max-width: 860px;
    }
    .status-strip {
        display: grid;
        grid-template-columns: repeat(4, minmax(0, 1fr));
        gap: 0.7rem;
        margin-top: 1rem;
    }
    .status-card {
        background: rgba(255,255,255,0.035);
        border: 1px solid rgba(212, 175, 55, 0.15);
        border-radius: 16px;
        padding: 0.9rem 1rem;
    }
    .status-label {
        color: #9FA9BA;
        text-transform: uppercase;
        font-size: 0.72rem;
        letter-spacing: 0.10em;
        margin-bottom: 0.3rem;
    }
    .status-value {
        color: #F5F1E6;
        font-size: 1.05rem;
        font-weight: 700;
    }
    .notice-panel, .mission-panel {
        border-radius: 18px;
        padding: 1rem 1.05rem;
        border: 1px solid rgba(212,175,55,0.16);
        background: rgba(255,255,255,0.035);
        height: 100%;
    }
    .panel-title {
        color: #D4AF37;
        text-transform: uppercase;
        letter-spacing: 0.12em;
        font-size: 0.76rem;
        font-weight: 700;
        margin-bottom: 0.35rem;
    }
    .panel-body {
        color: #D7DDEA;
        font-size: 0.93rem;
        line-height: 1.45;
    }
    .metric-shell {
        background: linear-gradient(180deg, rgba(14,20,31,0.95), rgba(9,12,18,0.98));
        border: 1px solid rgba(212,175,55,0.12);
        border-radius: 18px;
        padding: 0.2rem 0.35rem;
        box-shadow: 0 10px 26px rgba(0,0,0,0.18);
    }
    div[data-testid="metric-container"] {
        background: transparent;
        border: none;
        box-shadow: none;
        padding: 0.55rem 0.75rem 0.45rem 0.75rem;
    }
    div[data-testid="metric-container"] label {
        color: #A9B3C4 !important;
        text-transform: uppercase;
        letter-spacing: 0.08em;
    }
    div[data-testid="metric-container"] [data-testid="stMetricValue"] {
        color: #F7F2E7 !important;
    }
    .section-shell {
        border: 1px solid rgba(212,175,55,0.12);
        background: rgba(255,255,255,0.03);
        border-radius: 18px;
        padding: 1rem 1rem 0.6rem 1rem;
        margin-bottom: 1rem;
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 0.65rem;
    }
    .stTabs [data-baseweb="tab"] {
        background: rgba(255,255,255,0.03);
        border: 1px solid rgba(212,175,55,0.10);
        border-radius: 12px;
        padding: 0.6rem 1rem;
        color: #D5DDEB;
    }
    .stTabs [aria-selected="true"] {
        background: rgba(212,175,55,0.12) !important;
        color: #F5E7BC !important;
    }
    .small-callout {
        font-size: 0.85rem;
        color: #AAB4C5;
    }
    .footer-note {
        color: #9CA7B9;
        font-size: 0.83rem;
        margin-top: 0.8rem;
    }
    @media (max-width: 900px) {
        .hero-title { font-size: 2rem; }
        .status-strip { grid-template-columns: 1fr 1fr; }
    }


    /* V9 Export / widget contrast hotfix */
    div.stDownloadButton > button,
    button[kind="secondary"],
    button[data-testid="baseButton-secondary"] {
        background: linear-gradient(180deg, #182133 0%, #0F1724 100%) !important;
        color: #F4F1E8 !important;
        border: 1px solid rgba(212,175,55,0.45) !important;
        border-radius: 12px !important;
        font-weight: 700 !important;
        box-shadow: 0 0 0 1px rgba(255,255,255,0.03) inset, 0 8px 18px rgba(0,0,0,0.25) !important;
    }
    div.stDownloadButton > button:hover,
    button[kind="secondary"]:hover,
    button[data-testid="baseButton-secondary"]:hover {
        background: linear-gradient(180deg, rgba(212,175,55,0.22) 0%, #111827 100%) !important;
        color: #FFFFFF !important;
        border-color: #D4AF37 !important;
    }
    div[role="radiogroup"] label,
    div[role="radiogroup"] label span,
    div[data-testid="stRadio"] label,
    div[data-testid="stRadio"] p {
        color: #F4F1E8 !important;
        opacity: 1 !important;
    }
    div[data-baseweb="radio"] > div:first-child {
        border-color: rgba(212,175,55,0.65) !important;
    }
    div[data-baseweb="select"] > div,
    div[data-baseweb="select"] input,
    div[data-testid="stMultiSelect"] div[data-baseweb="select"] > div {
        background-color: #0F1724 !important;
        color: #F4F1E8 !important;
        border-color: rgba(212,175,55,0.35) !important;
    }
    div[data-baseweb="tag"] {
        background-color: rgba(212,175,55,0.18) !important;
        color: #F4F1E8 !important;
        border: 1px solid rgba(212,175,55,0.35) !important;
    }
    div[data-baseweb="tag"] span {
        color: #F4F1E8 !important;
    }
    label, .st-emotion-cache-ue6h4q, .st-emotion-cache-1v0mbdj, .st-emotion-cache-16idsys p {
        color: #D7DDEA !important;
        opacity: 1 !important;
    }

    </style>
    """, unsafe_allow_html=True)


def signal_from_hit_pct(value: float | None) -> str:
    if value is None or pd.isna(value):
        return "Insufficient data"
    if value >= 75:
        return "High alignment"
    if value >= 45:
        return "Moderate alignment"
    return "Low alignment"


def render_hero(datasets: list[Dataset]):
    st.markdown(f"""
    <div class="hero-shell">
        <div class="hero-kicker">Area 76 Command Center</div>
        <div class="hero-title">NONMS Genetics Engine</div>
        <div class="hero-sub">
            Upload AncestryDNA, 23andMe, or MyHeritage raw data and compare it against every modular SNP panel in /data.
            This interface is designed as a clean command-center view: signal first, detail second.
        </div>
        <div class="status-strip">
            <div class="status-card">
                <div class="status-label">Datasets bundled</div>
                <div class="status-value">{len(datasets)}</div>
            </div>
            <div class="status-card">
                <div class="status-label">Core layer</div>
                <div class="status-value">MS GWAS + pathways</div>
            </div>
            <div class="status-card">
                <div class="status-label">Export formats</div>
                <div class="status-value">PDF • Excel • CSV</div>
            </div>
            <div class="status-card">
                <div class="status-label">Mode</div>
                <div class="status-value">Pattern comparison</div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)


def render_top_panels():
    col1, col2 = st.columns([1.08, 1], gap="large")
    with col1:
        st.markdown("""
        <div class="notice-panel">
            <div class="panel-title">Privacy Notice</div>
            <div class="panel-body">
                Uploaded DNA files are processed in-session for comparison and report generation. This tool is intended for
                research and pattern exploration. It is not designed as a diagnosis, clinical interpretation, or medical advice.
            </div>
        </div>
        """, unsafe_allow_html=True)
    with col2:
        st.markdown("""
        <div class="mission-panel">
            <div class="panel-title">Mission Guardrails</div>
            <div class="panel-body">
                Treat outputs as signal mapping, not certainty. A marker match can support pattern exploration, but it does not
                prove disease, rule disease out, or replace clinical review.
            </div>
        </div>
        """, unsafe_allow_html=True)


def main():
    st.set_page_config(page_title="NONMS Genetics Engine", page_icon="🧬", layout="wide")
    inject_css()

    datasets = load_datasets_cached()
    render_hero(datasets)
    render_top_panels()

    with st.expander("What is included in Version 9?", expanded=False):
        st.write(", ".join(d.category for d in datasets))
        st.info("This modular app discovers every .txt SNP panel in /data. It performs literal marker comparison only; it does not diagnose disease or produce a validated risk score.")

    st.markdown('<div class="section-shell">', unsafe_allow_html=True)
    left, right = st.columns([1.25, 0.75], gap="large")
    with left:
        uploaded = st.file_uploader("Upload raw DNA file", type=["txt", "csv"], help="Supports AncestryDNA, 23andMe, and MyHeritage raw data exports.")
    with right:
        st.markdown("""
        <div class="small-callout">
        <strong>Recommended workflow</strong><br>
        1. Upload raw file<br>
        2. Review category summary<br>
        3. Inspect matched rows and manual-review rows<br>
        4. Export PDF or Excel report
        </div>
        """, unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

    if not uploaded:
        st.stop()

    ancestry_df = parse_dna_file(uploaded)
    st.success(f"Detected file type: {ancestry_df['source'].iloc[0]}")
    if ancestry_df.empty:
        st.error("No genotype rows could be parsed from the uploaded file.")
        st.stop()

    all_categories = [d.category for d in datasets]
    default_core = ["Immunometabolic Core", "MPOA Network", "Evolutionary Immune Network", "ERAP2 / Ancient Selection", "DIO Thermoregulation", "MS GWAS", "Methylation", "Dysautonomia"]
    default_selection = [c for c in default_core if c in all_categories] or all_categories

    st.markdown('<div class="section-shell">', unsafe_allow_html=True)
    selected_categories = st.multiselect(
        "Datasets to include",
        options=all_categories,
        default=default_selection,
        help="Choose the bundled signal layers you want included in this comparison run."
    )
    st.markdown('</div>', unsafe_allow_html=True)

    datasets = [d for d in datasets if d.category in selected_categories]
    result_frames = [compare_dataset(d, ancestry_df) for d in datasets]
    all_results = pd.concat(result_frames, ignore_index=True) if result_frames else pd.DataFrame()
    summary_df = build_summary(all_results) if not all_results.empty else pd.DataFrame()
    metrics = build_overall_metrics(summary_df, all_results) if not all_results.empty else {}

    if not metrics:
        st.warning("No datasets were selected.")
        st.stop()

    metric_cols = st.columns(6, gap="small")
    for i, (label, value) in enumerate(metrics.items()):
        with metric_cols[i]:
            st.markdown('<div class="metric-shell">', unsafe_allow_html=True)
            st.metric(label, value)
            st.markdown('</div>', unsafe_allow_html=True)

    if not summary_df.empty:
        summary_df = summary_df.copy()
        summary_df["signal_call"] = summary_df["match_pct_when_present"].apply(signal_from_hit_pct)

    tab_summary, tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "Volunteer Summary",
        "Command Summary",
        "Category Detail",
        "Matched Rows",
        "Manual Review",
        "Exports",
    ])


    with tab_summary:
        st.markdown('<div class="section-shell">', unsafe_allow_html=True)
        st.subheader("Plain-English Volunteer Summary")
        st.write("This section explains the report at a 7th-grade reading level. The detailed technical data is still available in the tabs that follow.")

        total_datasets = int(summary_df["category"].nunique()) if not summary_df.empty else 0
        total_markers = int(len(all_results)) if not all_results.empty else 0
        total_present = int(summary_df["present_in_ancestry"].sum()) if not summary_df.empty else 0

        st.markdown(f"""
        <div class="notice-panel">
            <div class="panel-title">What this report is</div>
            <div class="panel-body">
            This report reviewed <strong>{total_datasets}</strong> research panels containing <strong>{total_markers}</strong> marker rows.
            Your uploaded DNA file contained data for <strong>{total_present}</strong> of those marker rows.
            This is a research and education report, not a diagnosis.
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("### How to read the results")
        st.write("A marker match means the uploaded DNA file contained that SNP or listed allele. It does not prove disease. A missing marker usually means the DNA company did not test that SNP.")
        st.write("NONMS looks at pathways — groups of genes that may work together — rather than treating one gene as the whole answer.")

        st.markdown("### Your genetic story")
        for cat in get_top_categories_for_summary(summary_df, max_items=8):
            info = PATHWAY_PLAIN_ENGLISH.get(cat, {
                "plain": "This panel is one of the research pathways included in the NONMS genetics engine.",
                "why": "It is included to help researchers compare patterns across volunteers."
            })
            row = summary_df[summary_df["category"] == cat].iloc[0]
            hits = int(row["listed_allele_present"] + row["marker_present_no_allele"])
            present = int(row["present_in_ancestry"])
            rows_total = int(row["rows"])
            st.markdown(f"**{cat}**")
            st.write(info["plain"])
            st.caption(f"Research snapshot: {present} of {rows_total} marker rows were present; {hits} positive marker signal(s) were found. {info['why']}")
        st.markdown('</div>', unsafe_allow_html=True)


    with tab1:
        st.markdown('<div class="section-shell">', unsafe_allow_html=True)
        top_left, top_right = st.columns([1.1, 0.9], gap="large")
        with top_left:
            st.subheader("Signal board")
            st.dataframe(
                summary_df[[
                    "category", "rows", "directly_comparable_rows", "present_in_ancestry",
                    "listed_allele_present", "marker_present_no_allele",
                    "manual_review_rows", "coverage_pct_of_comparable",
                    "match_pct_when_present", "signal_call"
                ]],
                use_container_width=True,
                hide_index=True,
            )
        with top_right:
            st.subheader("Presence vs missing")
            chart_df = summary_df.set_index("category")[["present_in_ancestry", "missing_from_ancestry", "manual_review_rows"]]
            st.bar_chart(chart_df, use_container_width=True)
            st.caption("Use this to see where the volunteer file covered the bundled markers and where the DNA array left gaps.")
        st.markdown('</div>', unsafe_allow_html=True)

    with tab2:
        st.markdown('<div class="section-shell">', unsafe_allow_html=True)
        categories = summary_df["category"].tolist() if not summary_df.empty else []
        chosen = st.selectbox("Choose a category", categories) if categories else None
        if chosen:
            cat_df = all_results[all_results["category"] == chosen].copy()
            cat_summary = summary_df[summary_df["category"] == chosen].iloc[0]
            a, b, c = st.columns(3)
            a.metric("Signal call", cat_summary["signal_call"])
            b.metric("Coverage %", "-" if pd.isna(cat_summary["coverage_pct_of_comparable"]) else f"{cat_summary['coverage_pct_of_comparable']:.1f}%")
            c.metric("Match % when present", "-" if pd.isna(cat_summary["match_pct_when_present"]) else f"{cat_summary['match_pct_when_present']:.1f}%")
            st.dataframe(cat_df, use_container_width=True, hide_index=True)
            st.caption("Focus first on comparison_status, genotype, listed_allele, and manual_review.")
        st.markdown('</div>', unsafe_allow_html=True)

    with tab3:
        st.markdown('<div class="section-shell">', unsafe_allow_html=True)
        matched_df = all_results[all_results["comparison_status"].isin(["Listed allele present", "Marker present", "Possible exact multibase match"])].copy()
        st.subheader("Rows with direct positive signal")
        st.dataframe(matched_df, use_container_width=True, hide_index=True)
        st.markdown('</div>', unsafe_allow_html=True)

    with tab4:
        st.markdown('<div class="section-shell">', unsafe_allow_html=True)
        manual_df = all_results[all_results["manual_review"] != "No"].copy()
        st.subheader("Rows needing manual review")
        st.dataframe(manual_df, use_container_width=True, hide_index=True)
        st.caption("These rows typically involve HLA entries, composite markers, unknown alleles, or markers not directly testable from the uploaded raw file.")
        st.markdown('</div>', unsafe_allow_html=True)

    with tab5:
        st.markdown('<div class="section-shell">', unsafe_allow_html=True)
        st.subheader("Export report files")
        report_style = st.radio(
            "PDF report style",
            ["Print-Friendly (light)", "Command Center (dark)"],
            horizontal=True,
            help="Use the light version for printing and the dark version for on-screen presentation."
        )

        dl1, dl2, dl3, dl4 = st.columns(4)

        with dl1:
            try:
                excel_bytes = make_excel_report(summary_df, all_results, ancestry_df)
                st.download_button(
                    "Download Excel report (.xlsx)",
                    data=excel_bytes,
                    file_name="NONMS_Genetics_Report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as e:
                st.error(f"Excel export hit an error: {e}")

        with dl2:
            try:
                pdf_bytes = make_pdf_report(summary_df, all_results, uploaded.name, report_style=report_style)
                file_name = "NONMS_Genetics_Report_Print.pdf" if report_style.startswith("Print-Friendly") else "NONMS_Genetics_Report_CommandCenter.pdf"
                label = "Download selected PDF (.pdf)"
                st.download_button(
                    label,
                    data=pdf_bytes,
                    file_name=file_name,
                    mime="application/pdf",
                    use_container_width=True,
                )
            except Exception as e:
                st.error(f"PDF export hit an error: {e}")

        with dl3:
            try:
                pdf_light = make_pdf_report(summary_df, all_results, uploaded.name, report_style="Print-Friendly (light)")
                st.download_button(
                    "Download print PDF (.pdf)",
                    data=pdf_light,
                    file_name="NONMS_Genetics_Report_Print.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                )
            except Exception as e:
                st.error(f"Print PDF export hit an error: {e}")

        with dl4:
            try:
                csv_zip_bytes = make_csv_zip(summary_df, all_results)
                st.download_button(
                    "Download CSV bundle (.zip)",
                    data=csv_zip_bytes,
                    file_name="NONMS_Genetics_CSVs.zip",
                    mime="application/zip",
                    use_container_width=True,
                )
            except Exception as e:
                st.error(f"CSV export hit an error: {e}")

        st.markdown('<div class="footer-note">Use the print-friendly PDF for paper copies and doctor handouts. Use the dark command-center PDF for screen sharing, presentations, and brand-forward storytelling.</div>', unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)


if __name__ == "__main__":
    main()