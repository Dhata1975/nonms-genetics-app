
from __future__ import annotations

import io
import re
import zipfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional
from collections import defaultdict

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether
)

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"

DATASET_FILES = [
    "MS Attributes.txt",
    "SAM E Vulnerability.txt",
    "Stress Event.txt",
    "Small Vessel Disorder.txt",
    "Autonomic Loop.txt",
    "Molecular Mimicry.txt",
    "CSVD.txt",
    "Homocysteine.txt",
    "Mold Fungus.txt",
    "Tinea Versicolor.txt",
    "H Pylori.txt",
    "Cardiomegaly.txt",
    "Inverted T-waves.txt",
    "Low B12.txt",
    "Periodontal Disease.txt",
    "Methylation.txt",
    "Dysautonomia.txt",
]

DISPLAY_NAMES = {
    "MS Attributes.txt": "MS GWAS",
    "SAM E Vulnerability.txt": "SAM-e Vulnerability",
    "Stress Event.txt": "Stress Nexus Event",
    "Small Vessel Disorder.txt": "Small Vessel Disorder",
    "Autonomic Loop.txt": "Autonomic Loop",
    "Molecular Mimicry.txt": "Molecular Mimicry",
    "CSVD.txt": "CSVD",
    "Homocysteine.txt": "Homocysteine",
    "Mold Fungus.txt": "Mold / Fungus",
    "Tinea Versicolor.txt": "Tinea Versicolor",
    "H Pylori.txt": "H. pylori",
    "Cardiomegaly.txt": "Cardiomegaly",
    "Inverted T-waves.txt": "T-Waves",
    "Low B12.txt": "Low B12",
    "Periodontal Disease.txt": "Periodontal Disease",
    "Methylation.txt": "Methylation",
    "Dysautonomia.txt": "Dysautonomia",
}


@dataclass
class Dataset:
    file_name: str
    category: str
    frame: pd.DataFrame


def normalize_category_name(name: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "_", name).strip("_")[:31] or "Sheet"


def make_unique_sheet_name(base: str, used_names: set[str]) -> str:
    base = (base or "Sheet")[:31]
    if base not in used_names:
        used_names.add(base)
        return base
    i = 2
    while True:
        suffix = f"_{i}"
        candidate = f"{base[:31-len(suffix)]}{suffix}"
        if candidate not in used_names:
            used_names.add(candidate)
            return candidate
        i += 1


def clean_excel_value(value):
    if value is None:
        return None
    if isinstance(value, (int, float, bool)):
        return value
    s = str(value)
    # Remove control characters that can break openpyxl
    s = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", s)
    return s


def parse_marker_token(token: str) -> tuple[Optional[str], Optional[str], str, str]:
    token = str(token).strip()
    if not token:
        return None, None, "blank", "Blank marker"

    if ";" in token:
        rsids = re.findall(r"rs\d+", token)
        if len(rsids) == 1:
            return rsids[0], None, "composite", "Composite marker; manual review needed"
        return None, None, "composite", "Composite marker / haplotype entry; manual review needed"

    if token.startswith("DRB") or token.startswith("A*") or "*" in token:
        return None, None, "hla", "HLA allele entry; not directly testable from standard Ancestry raw SNP export"

    m = re.match(r"^(rs\d+)-([A-Za-z\?]+)$", token)
    if m:
        rsid, allele = m.groups()
        allele = allele.upper()
        if allele == "?":
            return rsid, None, "rsid_unknown_allele", "Marker present but allele unspecified"
        if re.fullmatch(r"[ACGTID]+", allele) and len(allele) == 1:
            return rsid, allele, "rsid_single_allele", "Allele-aware comparison available"
        return rsid, allele, "rsid_multibase_allele", "Non-single-base allele; compare with caution"

    m = re.match(r"^(rs\d+)$", token)
    if m:
        return m.group(1), None, "rsid_only", "Presence / absence comparison available"

    if token.startswith("chr"):
        return None, None, "coordinate_marker", "Coordinate-based marker; not directly matched because raw file is rsID-based"

    return None, None, "other", "Unrecognized marker format"


def normalize_dataset_lines(path: Path) -> list[str]:
    raw_lines = path.read_text(encoding="utf-8", errors="ignore").splitlines()
    cleaned = []
    i = 0
    marker_hint = re.compile(r"(\s|\t)(rs\S+|DRB\S+|A\*\S+|chr\S+|kgp\S+)$")
    while i < len(raw_lines):
        line = raw_lines[i].strip()
        if not line:
            i += 1
            continue
        if marker_hint.search(line) or line.startswith("#") or "rsID/SNP" in line or "TRAIT" in line:
            cleaned.append(line)
            i += 1
            continue
        if i + 1 < len(raw_lines):
            nxt = raw_lines[i + 1].strip()
            if marker_hint.search(nxt):
                cleaned.append(f"{line} {nxt}")
                i += 2
                continue
        cleaned.append(line)
        i += 1
    return cleaned


def parse_generic_dataset(path: Path) -> Dataset:
    rows = []
    fixed_category = DISPLAY_NAMES.get(path.name, path.stem)
    for line in normalize_dataset_lines(path):
        line = line.strip()
        if not line:
            continue
        if line.startswith("#") or "rsID/SNP" in line or "TRAIT" in line:
            continue

        entry_id = None
        trait_label = fixed_category
        marker = None

        parts = re.split(r"\t+", line)

        if len(parts) >= 3 and parts[0].strip().isdigit():
            entry_id = int(parts[0].strip())
            trait_label = parts[1].strip() or fixed_category
            marker = parts[2].strip()
        else:
            m = re.match(r"^(\d+)\s+(.+?)\s+(rs\S+|DRB\S+|A\*\S+|chr\S+|kgp\S+)$", line)
            if m:
                entry_id = int(m.group(1))
                trait_label = m.group(2).strip() or fixed_category
                marker = m.group(3).strip()
            else:
                if len(parts) >= 2:
                    trait_label = parts[0].strip() or fixed_category
                    marker = parts[1].strip()
                else:
                    marker = line.strip()

        rsid, allele, marker_type, note = parse_marker_token(marker)
        rows.append({
            "entry_id": entry_id,
            "category": fixed_category,
            "trait_label": trait_label,
            "raw_marker": marker,
            "rsid": rsid,
            "listed_allele": allele,
            "marker_type": marker_type,
            "note": note,
        })

    df = pd.DataFrame(rows)
    if df.empty:
        df = pd.DataFrame(columns=["entry_id", "category", "trait_label", "raw_marker", "rsid", "listed_allele", "marker_type", "note"])
    return Dataset(path.name, fixed_category, df)


def load_all_datasets() -> list[Dataset]:
    return [parse_generic_dataset(DATA_DIR / name) for name in DATASET_FILES if (DATA_DIR / name).exists()]


def parse_ancestry_file(uploaded_file) -> pd.DataFrame:
    rows = []
    content = uploaded_file.getvalue().decode("utf-8", errors="ignore").splitlines()
    for line in content:
        if not line or line.startswith("#") or line.lower().startswith("rsid\t"):
            continue
        parts = line.split("\t")
        if len(parts) != 5:
            continue
        rsid, chrom, pos, a1, a2 = parts
        try:
            pos = int(pos)
        except ValueError:
            continue
        rows.append((rsid.strip(), chrom.strip(), pos, a1.strip().upper(), a2.strip().upper()))
    return pd.DataFrame(rows, columns=["rsid", "chromosome", "position", "allele1", "allele2"])


def compare_dataset(dataset: Dataset, ancestry_df: pd.DataFrame) -> pd.DataFrame:
    df = dataset.frame.copy()
    if ancestry_df.empty:
        return df.assign(
            in_ancestry=False, chromosome=None, position=None, allele1=None, allele2=None,
            genotype=None, listed_allele_copies=None, zygosity=None,
            comparison_status="No ancestry data loaded", manual_review="Yes"
        )

    anc = ancestry_df.drop_duplicates("rsid").set_index("rsid")
    result_rows = []

    for _, row in df.iterrows():
        rsid = row["rsid"]
        allele = row["listed_allele"]
        marker_type = row["marker_type"]

        if not rsid:
            result_rows.append({
                **row.to_dict(),
                "in_ancestry": False,
                "chromosome": None,
                "position": None,
                "allele1": None,
                "allele2": None,
                "genotype": None,
                "listed_allele_copies": None,
                "zygosity": None,
                "comparison_status": "Not directly comparable",
                "manual_review": "Yes",
            })
            continue

        if rsid not in anc.index:
            result_rows.append({
                **row.to_dict(),
                "in_ancestry": False,
                "chromosome": None,
                "position": None,
                "allele1": None,
                "allele2": None,
                "genotype": None,
                "listed_allele_copies": None,
                "zygosity": None,
                "comparison_status": "rsID not present in Ancestry file",
                "manual_review": "Maybe",
            })
            continue

        rec = anc.loc[rsid]
        genotype = f"{rec['allele1']}{rec['allele2']}"

        status = ""
        copies = None
        zygosity = None
        manual_review = "No"

        if marker_type == "rsid_single_allele":
            copies = int(rec["allele1"] == allele) + int(rec["allele2"] == allele)
            if copies == 2:
                status = "Listed allele present"
                zygosity = "Homozygous listed allele"
            elif copies == 1:
                status = "Listed allele present"
                zygosity = "Heterozygous"
            else:
                status = "Listed allele absent"
                zygosity = "Listed allele absent"
        elif marker_type == "rsid_only":
            status = "Marker present"
            zygosity = "Genotype observed"
        elif marker_type == "rsid_unknown_allele":
            status = "Present, but listed allele unspecified"
            manual_review = "Yes"
        elif marker_type == "rsid_multibase_allele":
            copies = int(rec["allele1"] == allele) + int(rec["allele2"] == allele)
            status = "Possible exact multibase match" if copies > 0 else "No exact multibase match"
            zygosity = "Contains multibase allele" if copies > 0 else "Multibase allele not detected"
            manual_review = "Yes"
        else:
            status = "Present, manual review needed"
            manual_review = "Yes"

        result_rows.append({
            **row.to_dict(),
            "in_ancestry": True,
            "chromosome": rec["chromosome"],
            "position": int(rec["position"]),
            "allele1": rec["allele1"],
            "allele2": rec["allele2"],
            "genotype": genotype,
            "listed_allele_copies": copies,
            "zygosity": zygosity,
            "comparison_status": status,
            "manual_review": manual_review,
        })

    return pd.DataFrame(result_rows)


def build_summary(all_results: pd.DataFrame) -> pd.DataFrame:
    summaries = []
    for category, grp in all_results.groupby("category", sort=True):
        total = len(grp)
        directly_comparable = int(grp["marker_type"].isin(["rsid_single_allele", "rsid_only"]).sum())
        found = int(grp["in_ancestry"].fillna(False).sum())
        allele_present = int((grp["comparison_status"] == "Listed allele present").sum())
        marker_present = int((grp["comparison_status"] == "Marker present").sum())
        absent = int((grp["comparison_status"] == "Listed allele absent").sum())
        missing = int((grp["comparison_status"] == "rsID not present in Ancestry file").sum())
        manual = int((grp["manual_review"] != "No").sum())
        comparable_found = int(((grp["marker_type"].isin(["rsid_single_allele", "rsid_only"])) & (grp["in_ancestry"].fillna(False))).sum())
        coverage_pct = round(100 * comparable_found / directly_comparable, 1) if directly_comparable else None
        hit_pct = round(100 * (allele_present + marker_present) / comparable_found, 1) if comparable_found else None

        summaries.append({
            "category": category,
            "rows": total,
            "directly_comparable_rows": directly_comparable,
            "present_in_ancestry": found,
            "listed_allele_present": allele_present,
            "marker_present_no_allele": marker_present,
            "listed_allele_absent": absent,
            "missing_from_ancestry": missing,
            "manual_review_rows": manual,
            "coverage_pct_of_comparable": coverage_pct,
            "match_pct_when_present": hit_pct,
        })

    return pd.DataFrame(summaries).sort_values(["category"]).reset_index(drop=True)


def build_overall_metrics(summary_df: pd.DataFrame, all_results: pd.DataFrame) -> dict:
    if summary_df.empty:
        return {}
    return {
        "Datasets loaded": int(summary_df["category"].nunique()),
        "Total marker rows": int(len(all_results)),
        "Directly comparable rows": int(summary_df["directly_comparable_rows"].sum()),
        "Present in volunteer file": int(summary_df["present_in_ancestry"].sum()),
        "Allele / marker hits": int((summary_df["listed_allele_present"] + summary_df["marker_present_no_allele"]).sum()),
        "Manual review rows": int(summary_df["manual_review_rows"].sum()),
    }


def make_excel_report(summary_df: pd.DataFrame, all_results: pd.DataFrame, ancestry_df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "README"
    used_names = {"README"}

    navy = PatternFill("solid", fgColor="13263A")
    white_bold = Font(color="FFFFFF", bold=True)

    ws0.merge_cells("A1:F1")
    ws0["A1"] = "NONMS Genetics Comparison Report"
    ws0["A1"].fill = navy
    ws0["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    ws0["A3"] = "Generated"
    ws0["B3"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    ws0["A5"] = "Important"
    ws0["B5"] = "This workbook is a pattern-comparison tool. It is not a diagnosis, clinical interpretation, or validated polygenic risk score."
    ws0["A7"] = "Input file rows"
    ws0["B7"] = len(ancestry_df)
    ws0["A8"] = "Marker rows reviewed"
    ws0["B8"] = len(all_results)
    ws0.column_dimensions["A"].width = 22
    ws0.column_dimensions["B"].width = 100
    ws0.sheet_view.showGridLines = False

    def add_df_sheet(name: str, df: pd.DataFrame):
        safe_name = make_unique_sheet_name(name[:31], used_names)
        ws = wb.create_sheet(safe_name)
        if df.empty:
            ws["A1"] = "No rows"
            return ws
        for c_idx, col in enumerate(df.columns, 1):
            cell = ws.cell(1, c_idx, clean_excel_value(col))
            cell.fill = navy
            cell.font = white_bold
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for r_idx, row in enumerate(df.itertuples(index=False), 2):
            for c_idx, val in enumerate(row, 1):
                ws.cell(r_idx, c_idx, clean_excel_value(val))
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(df.shape[1])}{ws.max_row}"
        ws.sheet_view.showGridLines = False
        for idx, col in enumerate(df.columns, start=1):
            width = max(len(str(col)) + 2, 12)
            sample_vals = [clean_excel_value(v) for v in df[col].head(200).tolist()]
            sample_vals = ["" if v is None else str(v) for v in sample_vals]
            if sample_vals:
                width = max(width, min(42, int(max(len(v) for v in sample_vals) * 0.95) + 2))
            ws.column_dimensions[get_column_letter(idx)].width = min(max(width, 12), 45)
        return ws

    add_df_sheet("Category_Summary", summary_df)
    add_df_sheet("All_Results", all_results)
    add_df_sheet("Matched_Only", all_results[all_results["comparison_status"].isin(["Listed allele present", "Marker present", "Possible exact multibase match"])])
    add_df_sheet("Manual_Review", all_results[all_results["manual_review"] != "No"])

    for category in summary_df["category"].tolist():
        cat_df = all_results[all_results["category"] == category].copy()
        add_df_sheet(normalize_category_name(category), cat_df)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()



def draw_pdf_header(canvas, doc):
    canvas.saveState()
    page_w, page_h = doc.pagesize

    canvas.setFillColor(colors.HexColor("#070A0F"))
    canvas.rect(0, 0, page_w, page_h, fill=1, stroke=0)

    canvas.setFillColor(colors.HexColor("#0F1724"))
    canvas.rect(0.45 * inch, page_h - 0.95 * inch, page_w - 0.9 * inch, 0.42 * inch, fill=1, stroke=0)

    canvas.setFillColor(colors.HexColor("#D4AF37"))
    canvas.rect(0.45 * inch, page_h - 0.53 * inch, page_w - 0.9 * inch, 0.05 * inch, fill=1, stroke=0)

    canvas.setFillColor(colors.HexColor("#8FA7C6"))
    canvas.setFont("Helvetica-Bold", 9)
    canvas.drawString(0.62 * inch, page_h - 0.79 * inch, "AREA 76 // NONMS COMMAND CENTER")

    canvas.setFillColor(colors.white)
    canvas.setFont("Helvetica-Bold", 9)
    canvas.drawRightString(page_w - 0.62 * inch, page_h - 0.79 * inch, "GENETICS SIGNAL REPORT")

    canvas.setFillColor(colors.HexColor("#7E8798"))
    canvas.setFont("Helvetica", 8.5)
    canvas.drawString(0.62 * inch, 0.38 * inch, "Pattern comparison only. Not a diagnosis or medical advice.")
    canvas.drawRightString(page_w - 0.62 * inch, 0.38 * inch, f"Page {doc.page}")
    canvas.restoreState()


def _metric_card(label: str, value: str, width: float):
    data = [[label], [value]]
    t = Table(data, colWidths=[width])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#13263A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#9FB6D5")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#10151E")),
        ("TEXTCOLOR", (0, 1), (-1, 1), colors.HexColor("#F5E8B1")),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 1), (-1, 1), 15),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#D4AF37")),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    return t


def _section_table(data, col_widths):
    table = Table(data, repeatRows=1, colWidths=col_widths)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D4AF37")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#10151E")),
        ("TEXTCOLOR", (0, 1), (-1, -1), colors.white),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#10151E"), colors.HexColor("#0C1118")]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#505A69")),
        ("FONTSIZE", (0, 0), (-1, -1), 8.2),
        ("LEADING", (0, 0), (-1, -1), 10.5),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    return table


def make_pdf_report(summary_df: pd.DataFrame, all_results: pd.DataFrame, volunteer_filename: str) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=0.62 * inch,
        rightMargin=0.62 * inch,
        topMargin=1.05 * inch,
        bottomMargin=0.65 * inch,
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="A76_Kicker",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=9,
        leading=11,
        textColor=colors.HexColor("#8FA7C6"),
        spaceAfter=4,
    ))
    styles.add(ParagraphStyle(
        name="A76_Title",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=24,
        leading=28,
        textColor=colors.HexColor("#F5E8B1"),
        spaceAfter=6,
    ))
    styles.add(ParagraphStyle(
        name="A76_Subtitle",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=10,
        leading=14,
        textColor=colors.white,
        spaceAfter=8,
    ))
    styles.add(ParagraphStyle(
        name="A76_Label",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=11.5,
        leading=14,
        textColor=colors.HexColor("#D4AF37"),
        spaceBefore=8,
        spaceAfter=6,
    ))
    styles.add(ParagraphStyle(
        name="A76_Body",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=9.4,
        leading=13,
        textColor=colors.white,
    ))
    styles.add(ParagraphStyle(
        name="A76_Small",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8.4,
        leading=11,
        textColor=colors.HexColor("#C8D0DE"),
    ))

    story = []

    # Cover / executive page
    story.append(Paragraph("AREA 76 // NONMS COMMAND CENTER", styles["A76_Kicker"]))
    story.append(Paragraph("GENETICS SIGNAL REPORT", styles["A76_Title"]))
    story.append(Paragraph(
        "Volunteer upload processed through bundled MS and pathway marker sets. "
        "This report is structured as a command summary: high-level signal first, drill-down second.",
        styles["A76_Subtitle"],
    ))

    meta_table = _section_table(
        [
            ["Volunteer File", "Generated (UTC)", "Report Mode"],
            [clean_excel_value(volunteer_filename), datetime.utcnow().strftime("%Y-%m-%d %H:%M"), "Pattern comparison"],
        ],
        [2.85 * inch, 2.0 * inch, 1.6 * inch],
    )
    story.append(meta_table)
    story.append(Spacer(1, 0.16 * inch))

    total_rows = len(all_results)
    comparable = int(summary_df["directly_comparable_rows"].sum()) if not summary_df.empty else 0
    present = int(summary_df["present_in_ancestry"].sum()) if not summary_df.empty else 0
    hits = int((summary_df["listed_allele_present"] + summary_df["marker_present_no_allele"]).sum()) if not summary_df.empty else 0

    cards = Table([[
        _metric_card("Datasets", str(summary_df["category"].nunique() if not summary_df.empty else 0), 1.45 * inch),
        _metric_card("Marker rows", str(total_rows), 1.55 * inch),
        _metric_card("Comparable", str(comparable), 1.45 * inch),
        _metric_card("Present in file", str(present), 1.55 * inch),
        _metric_card("Positive hits", str(hits), 1.45 * inch),
    ]], colWidths=[1.45 * inch, 1.55 * inch, 1.45 * inch, 1.55 * inch, 1.45 * inch])
    cards.setStyle(TableStyle([("VALIGN", (0,0), (-1,-1), "TOP")]))
    story.append(cards)
    story.append(Spacer(1, 0.18 * inch))

    story.append(Paragraph("Mission Guardrails", styles["A76_Label"]))
    story.append(Paragraph(
        "Treat every match here as a pattern signal, not a conclusion. Marker presence can support hypothesis generation "
        "and comparison across volunteers, but it does not prove disease, rule disease out, or replace clinical interpretation.",
        styles["A76_Body"],
    ))
    story.append(Spacer(1, 0.14 * inch))

    top_summary = summary_df.copy()
    if not top_summary.empty:
        top_summary["signal_call"] = top_summary["match_pct_when_present"].apply(signal_from_hit_pct)
        top_summary = top_summary.sort_values(
            by=["match_pct_when_present", "present_in_ancestry", "rows"],
            ascending=[False, False, False],
            na_position="last",
        )

    story.append(Paragraph("Executive Signal Board", styles["A76_Label"]))
    summary_data = [[
        "Category", "Rows", "Present", "Hits", "Coverage %", "Match %", "Signal Call"
    ]]
    for _, row in top_summary.iterrows():
        summary_data.append([
            str(row["category"]),
            int(row["rows"]),
            int(row["present_in_ancestry"]),
            int(row["listed_allele_present"] + row["marker_present_no_allele"]),
            "-" if pd.isna(row["coverage_pct_of_comparable"]) else f'{row["coverage_pct_of_comparable"]:.1f}',
            "-" if pd.isna(row["match_pct_when_present"]) else f'{row["match_pct_when_present"]:.1f}',
            str(row["signal_call"]),
        ])
    story.append(_section_table(summary_data, [2.1 * inch, 0.55 * inch, 0.7 * inch, 0.55 * inch, 0.8 * inch, 0.72 * inch, 1.15 * inch]))
    story.append(Spacer(1, 0.15 * inch))

    top_hits = all_results[all_results["comparison_status"].isin(["Listed allele present", "Marker present", "Possible exact multibase match"])].copy()
    story.append(Paragraph("Command Readout", styles["A76_Label"]))
    if not top_hits.empty:
        by_cat = top_hits.groupby("category").size().sort_values(ascending=False).head(6)
        readout_lines = [f"<b>{cat}</b>: {count} positive row(s)" for cat, count in by_cat.items()]
        story.append(Paragraph(" • ".join(readout_lines), styles["A76_Body"]))
    else:
        story.append(Paragraph("No direct positive rows were detected in the selected categories.", styles["A76_Body"]))

    story.append(PageBreak())

    # Per-category pages
    for category in top_summary["category"].tolist():
        cat_summary = top_summary[top_summary["category"] == category].iloc[0]
        grp = all_results[all_results["category"] == category].copy()

        signal = str(cat_summary["signal_call"])
        coverage = "-" if pd.isna(cat_summary["coverage_pct_of_comparable"]) else f'{cat_summary["coverage_pct_of_comparable"]:.1f}%'
        match_pct = "-" if pd.isna(cat_summary["match_pct_when_present"]) else f'{cat_summary["match_pct_when_present"]:.1f}%'

        story.append(KeepTogether([
            Paragraph(category, styles["A76_Title"]),
            Paragraph(
                f"Signal call: <b>{signal}</b> &nbsp;&nbsp;|&nbsp;&nbsp; Coverage: <b>{coverage}</b> &nbsp;&nbsp;|&nbsp;&nbsp; Match when present: <b>{match_pct}</b>",
                styles["A76_Subtitle"],
            ),
        ]))

        mini = _section_table(
            [
                ["Rows", "Comparable", "Present", "Hits", "Missing", "Manual Review"],
                [
                    int(cat_summary["rows"]),
                    int(cat_summary["directly_comparable_rows"]),
                    int(cat_summary["present_in_ancestry"]),
                    int(cat_summary["listed_allele_present"] + cat_summary["marker_present_no_allele"]),
                    int(cat_summary["missing_from_ancestry"]),
                    int(cat_summary["manual_review_rows"]),
                ],
            ],
            [0.8 * inch, 1.0 * inch, 0.8 * inch, 0.7 * inch, 0.8 * inch, 1.05 * inch],
        )
        story.append(mini)
        story.append(Spacer(1, 0.12 * inch))

        positives = grp[grp["comparison_status"].isin(["Listed allele present", "Marker present", "Possible exact multibase match"])].copy().head(12)
        manual = grp[grp["manual_review"] != "No"].copy().head(10)

        story.append(Paragraph("Positive Signal Rows", styles["A76_Label"]))
        if positives.empty:
            story.append(Paragraph("No direct positive rows in this category for the current volunteer file.", styles["A76_Body"]))
        else:
            pos_data = [["Marker", "rsID", "Allele", "Genotype", "Status"]]
            for _, r in positives.iterrows():
                pos_data.append([
                    str(r["raw_marker"])[:30],
                    str(r["rsid"] or ""),
                    str(r["listed_allele"] or ""),
                    str(r["genotype"] or ""),
                    str(r["comparison_status"] or ""),
                ])
            story.append(_section_table(pos_data, [2.05 * inch, 1.1 * inch, 0.55 * inch, 0.8 * inch, 2.2 * inch]))

        story.append(Spacer(1, 0.12 * inch))
        story.append(Paragraph("Manual Review Flags", styles["A76_Label"]))
        if manual.empty:
            story.append(Paragraph("No manual-review rows surfaced in the top slice for this category.", styles["A76_Body"]))
        else:
            man_data = [["Marker", "rsID", "Issue", "Note"]]
            for _, r in manual.iterrows():
                man_data.append([
                    str(r["raw_marker"])[:28],
                    str(r["rsid"] or ""),
                    str(r["marker_type"] or ""),
                    str(r["note"] or "")[:52],
                ])
            story.append(_section_table(man_data, [1.95 * inch, 1.05 * inch, 1.25 * inch, 2.75 * inch]))

        story.append(Spacer(1, 0.12 * inch))
        story.append(Paragraph(
            "Analyst note: use the Excel export for complete row-level detail, especially when a category contains HLA entries, composite markers, or unknown alleles.",
            styles["A76_Small"],
        ))

        if category != top_summary["category"].tolist()[-1]:
            story.append(PageBreak())

    doc.build(story, onFirstPage=draw_pdf_header, onLaterPages=draw_pdf_header)
    return buffer.getvalue()


def make_csv_zip(summary_df: pd.DataFrame, all_results: pd.DataFrame) -> bytes:
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("category_summary.csv", summary_df.to_csv(index=False))
        zf.writestr("all_results.csv", all_results.to_csv(index=False))
        for category in summary_df["category"].tolist():
            cat_df = all_results[all_results["category"] == category]
            zf.writestr(f"{normalize_category_name(category)}.csv", cat_df.to_csv(index=False))
    return buffer.getvalue()


@st.cache_data
def load_datasets_cached():
    return load_all_datasets()


def inject_css():
    st.markdown("""
    <style>
    .stApp {
        background:
            radial-gradient(circle at top left, rgba(212, 175, 55, 0.09), transparent 28%),
            radial-gradient(circle at top right, rgba(62, 108, 168, 0.10), transparent 22%),
            linear-gradient(180deg, #06080D 0%, #0B0F17 45%, #090B10 100%);
        color: #F4F1E8;
    }
    .block-container {
        padding-top: 1.4rem;
        padding-bottom: 2rem;
        max-width: 1380px;
    }
    .hero-shell {
        border: 1px solid rgba(212, 175, 55, 0.20);
        background: linear-gradient(180deg, rgba(15,22,35,0.92), rgba(10,14,22,0.96));
        border-radius: 22px;
        padding: 1.4rem 1.5rem 1.15rem 1.5rem;
        box-shadow: 0 0 0 1px rgba(255,255,255,0.03) inset, 0 18px 44px rgba(0,0,0,0.35);
        margin-bottom: 1rem;
    }
    .hero-kicker {
        color: #D4AF37;
        font-size: 0.78rem;
        letter-spacing: 0.18em;
        text-transform: uppercase;
        font-weight: 700;
        margin-bottom: 0.35rem;
    }
    .hero-title {
        font-size: 2.55rem;
        line-height: 1.05;
        font-weight: 800;
        color: #F7F2E7;
        margin: 0 0 0.35rem 0;
    }
    .hero-sub {
        color: #B7C0D0;
        font-size: 1rem;
        margin-bottom: 1rem;
        max-width: 860px;
    }
    .status-strip {
        display: grid;
        grid-template-columns: repeat(4, minmax(0, 1fr));
        gap: 0.7rem;
        margin-top: 1rem;
    }
    .status-card {
        background: rgba(255,255,255,0.035);
        border: 1px solid rgba(212, 175, 55, 0.15);
        border-radius: 16px;
        padding: 0.9rem 1rem;
    }
    .status-label {
        color: #9FA9BA;
        text-transform: uppercase;
        font-size: 0.72rem;
        letter-spacing: 0.10em;
        margin-bottom: 0.3rem;
    }
    .status-value {
        color: #F5F1E6;
        font-size: 1.05rem;
        font-weight: 700;
    }
    .notice-panel, .mission-panel {
        border-radius: 18px;
        padding: 1rem 1.05rem;
        border: 1px solid rgba(212,175,55,0.16);
        background: rgba(255,255,255,0.035);
        height: 100%;
    }
    .panel-title {
        color: #D4AF37;
        text-transform: uppercase;
        letter-spacing: 0.12em;
        font-size: 0.76rem;
        font-weight: 700;
        margin-bottom: 0.35rem;
    }
    .panel-body {
        color: #D7DDEA;
        font-size: 0.93rem;
        line-height: 1.45;
    }
    .metric-shell {
        background: linear-gradient(180deg, rgba(14,20,31,0.95), rgba(9,12,18,0.98));
        border: 1px solid rgba(212,175,55,0.12);
        border-radius: 18px;
        padding: 0.2rem 0.35rem;
        box-shadow: 0 10px 26px rgba(0,0,0,0.18);
    }
    div[data-testid="metric-container"] {
        background: transparent;
        border: none;
        box-shadow: none;
        padding: 0.55rem 0.75rem 0.45rem 0.75rem;
    }
    div[data-testid="metric-container"] label {
        color: #A9B3C4 !important;
        text-transform: uppercase;
        letter-spacing: 0.08em;
    }
    div[data-testid="metric-container"] [data-testid="stMetricValue"] {
        color: #F7F2E7 !important;
    }
    .section-shell {
        border: 1px solid rgba(212,175,55,0.12);
        background: rgba(255,255,255,0.03);
        border-radius: 18px;
        padding: 1rem 1rem 0.6rem 1rem;
        margin-bottom: 1rem;
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 0.65rem;
    }
    .stTabs [data-baseweb="tab"] {
        background: rgba(255,255,255,0.03);
        border: 1px solid rgba(212,175,55,0.10);
        border-radius: 12px;
        padding: 0.6rem 1rem;
        color: #D5DDEB;
    }
    .stTabs [aria-selected="true"] {
        background: rgba(212,175,55,0.12) !important;
        color: #F5E7BC !important;
    }
    .small-callout {
        font-size: 0.85rem;
        color: #AAB4C5;
    }
    .footer-note {
        color: #9CA7B9;
        font-size: 0.83rem;
        margin-top: 0.8rem;
    }
    @media (max-width: 900px) {
        .hero-title { font-size: 2rem; }
        .status-strip { grid-template-columns: 1fr 1fr; }
    }
    </style>
    """, unsafe_allow_html=True)


def signal_from_hit_pct(value: float | None) -> str:
    if value is None or pd.isna(value):
        return "Insufficient data"
    if value >= 75:
        return "High alignment"
    if value >= 45:
        return "Moderate alignment"
    return "Low alignment"


def render_hero(datasets: list[Dataset]):
    st.markdown(f"""
    <div class="hero-shell">
        <div class="hero-kicker">Area 76 Command Center</div>
        <div class="hero-title">NONMS Genetics Engine</div>
        <div class="hero-sub">
            Upload an AncestryDNA raw file and run a structured comparison against bundled MS and biological pathway marker sets.
            This interface is designed as a clean command-center view: signal first, detail second.
        </div>
        <div class="status-strip">
            <div class="status-card">
                <div class="status-label">Datasets bundled</div>
                <div class="status-value">{len(datasets)}</div>
            </div>
            <div class="status-card">
                <div class="status-label">Core layer</div>
                <div class="status-value">MS GWAS + pathways</div>
            </div>
            <div class="status-card">
                <div class="status-label">Export formats</div>
                <div class="status-value">PDF • Excel • CSV</div>
            </div>
            <div class="status-card">
                <div class="status-label">Mode</div>
                <div class="status-value">Pattern comparison</div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)


def render_top_panels():
    col1, col2 = st.columns([1.08, 1], gap="large")
    with col1:
        st.markdown("""
        <div class="notice-panel">
            <div class="panel-title">Privacy Notice</div>
            <div class="panel-body">
                Uploaded DNA files are processed in-session for comparison and report generation. This tool is intended for
                research and pattern exploration. It is not designed as a diagnosis, clinical interpretation, or medical advice.
            </div>
        </div>
        """, unsafe_allow_html=True)
    with col2:
        st.markdown("""
        <div class="mission-panel">
            <div class="panel-title">Mission Guardrails</div>
            <div class="panel-body">
                Treat outputs as signal mapping, not certainty. A marker match can support pattern exploration, but it does not
                prove disease, rule disease out, or replace clinical review.
            </div>
        </div>
        """, unsafe_allow_html=True)


def main():
    st.set_page_config(page_title="NONMS Genetics Engine", page_icon="🧬", layout="wide")
    inject_css()

    datasets = load_datasets_cached()
    render_hero(datasets)
    render_top_panels()

    with st.expander("What is included in v4?", expanded=False):
        st.write(", ".join(DISPLAY_NAMES.get(name, name) for name in DATASET_FILES))
        st.info("This app performs literal marker comparison. It does not diagnose disease or produce a validated risk score.")

    st.markdown('<div class="section-shell">', unsafe_allow_html=True)
    left, right = st.columns([1.25, 0.75], gap="large")
    with left:
        uploaded = st.file_uploader("Upload AncestryDNA raw .txt file", type=["txt"], help="Use the raw data text file exported from AncestryDNA.")
    with right:
        st.markdown("""
        <div class="small-callout">
        <strong>Recommended workflow</strong><br>
        1. Upload raw file<br>
        2. Review category summary<br>
        3. Inspect matched rows and manual-review rows<br>
        4. Export PDF or Excel report
        </div>
        """, unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

    if not uploaded:
        st.stop()

    ancestry_df = parse_ancestry_file(uploaded)
    if ancestry_df.empty:
        st.error("No genotype rows could be parsed from the uploaded file.")
        st.stop()

    all_categories = [d.category for d in datasets]
    default_core = ["MS GWAS", "Methylation", "Mold / Fungus", "Autonomic Loop", "Molecular Mimicry", "Dysautonomia"]
    default_selection = [c for c in default_core if c in all_categories] or all_categories

    st.markdown('<div class="section-shell">', unsafe_allow_html=True)
    selected_categories = st.multiselect(
        "Datasets to include",
        options=all_categories,
        default=default_selection,
        help="Choose the bundled signal layers you want included in this comparison run."
    )
    st.markdown('</div>', unsafe_allow_html=True)

    datasets = [d for d in datasets if d.category in selected_categories]
    result_frames = [compare_dataset(d, ancestry_df) for d in datasets]
    all_results = pd.concat(result_frames, ignore_index=True) if result_frames else pd.DataFrame()
    summary_df = build_summary(all_results) if not all_results.empty else pd.DataFrame()
    metrics = build_overall_metrics(summary_df, all_results) if not all_results.empty else {}

    if not metrics:
        st.warning("No datasets were selected.")
        st.stop()

    metric_cols = st.columns(6, gap="small")
    for i, (label, value) in enumerate(metrics.items()):
        with metric_cols[i]:
            st.markdown('<div class="metric-shell">', unsafe_allow_html=True)
            st.metric(label, value)
            st.markdown('</div>', unsafe_allow_html=True)

    if not summary_df.empty:
        summary_df = summary_df.copy()
        summary_df["signal_call"] = summary_df["match_pct_when_present"].apply(signal_from_hit_pct)

    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "Command Summary",
        "Category Detail",
        "Matched Rows",
        "Manual Review",
        "Exports",
    ])

    with tab1:
        st.markdown('<div class="section-shell">', unsafe_allow_html=True)
        top_left, top_right = st.columns([1.1, 0.9], gap="large")
        with top_left:
            st.subheader("Signal board")
            st.dataframe(
                summary_df[[
                    "category", "rows", "directly_comparable_rows", "present_in_ancestry",
                    "listed_allele_present", "marker_present_no_allele",
                    "manual_review_rows", "coverage_pct_of_comparable",
                    "match_pct_when_present", "signal_call"
                ]],
                use_container_width=True,
                hide_index=True,
            )
        with top_right:
            st.subheader("Presence vs missing")
            chart_df = summary_df.set_index("category")[["present_in_ancestry", "missing_from_ancestry", "manual_review_rows"]]
            st.bar_chart(chart_df, use_container_width=True)
            st.caption("Use this to see where the volunteer file covered the bundled markers and where the Ancestry array left gaps.")
        st.markdown('</div>', unsafe_allow_html=True)

    with tab2:
        st.markdown('<div class="section-shell">', unsafe_allow_html=True)
        categories = summary_df["category"].tolist() if not summary_df.empty else []
        chosen = st.selectbox("Choose a category", categories) if categories else None
        if chosen:
            cat_df = all_results[all_results["category"] == chosen].copy()
            cat_summary = summary_df[summary_df["category"] == chosen].iloc[0]
            a, b, c = st.columns(3)
            a.metric("Signal call", cat_summary["signal_call"])
            b.metric("Coverage %", "-" if pd.isna(cat_summary["coverage_pct_of_comparable"]) else f"{cat_summary['coverage_pct_of_comparable']:.1f}%")
            c.metric("Match % when present", "-" if pd.isna(cat_summary["match_pct_when_present"]) else f"{cat_summary['match_pct_when_present']:.1f}%")
            st.dataframe(cat_df, use_container_width=True, hide_index=True)
            st.caption("Focus first on comparison_status, genotype, listed_allele, and manual_review.")
        st.markdown('</div>', unsafe_allow_html=True)

    with tab3:
        st.markdown('<div class="section-shell">', unsafe_allow_html=True)
        matched_df = all_results[all_results["comparison_status"].isin(["Listed allele present", "Marker present", "Possible exact multibase match"])].copy()
        st.subheader("Rows with direct positive signal")
        st.dataframe(matched_df, use_container_width=True, hide_index=True)
        st.markdown('</div>', unsafe_allow_html=True)

    with tab4:
        st.markdown('<div class="section-shell">', unsafe_allow_html=True)
        manual_df = all_results[all_results["manual_review"] != "No"].copy()
        st.subheader("Rows needing manual review")
        st.dataframe(manual_df, use_container_width=True, hide_index=True)
        st.caption("These rows typically involve HLA entries, composite markers, unknown alleles, or markers not directly testable from the uploaded raw file.")
        st.markdown('</div>', unsafe_allow_html=True)

    with tab5:
        st.markdown('<div class="section-shell">', unsafe_allow_html=True)
        st.subheader("Export report files")
        dl1, dl2, dl3 = st.columns(3)

        with dl1:
            try:
                excel_bytes = make_excel_report(summary_df, all_results, ancestry_df)
                st.download_button(
                    "Download Excel report (.xlsx)",
                    data=excel_bytes,
                    file_name="NONMS_Genetics_Report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as e:
                st.error(f"Excel export hit an error: {e}")

        with dl2:
            try:
                pdf_bytes = make_pdf_report(summary_df, all_results, uploaded.name)
                st.download_button(
                    "Download PDF summary (.pdf)",
                    data=pdf_bytes,
                    file_name="NONMS_Genetics_Report.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                )
            except Exception as e:
                st.error(f"PDF export hit an error: {e}")

        with dl3:
            try:
                csv_zip_bytes = make_csv_zip(summary_df, all_results)
                st.download_button(
                    "Download CSV bundle (.zip)",
                    data=csv_zip_bytes,
                    file_name="NONMS_Genetics_CSVs.zip",
                    mime="application/zip",
                    use_container_width=True,
                )
            except Exception as e:
                st.error(f"CSV export hit an error: {e}")

        st.markdown('<div class="footer-note">The PDF is a concise command summary. The Excel export contains full row-level detail by category. If one export fails, the others remain available.</div>', unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)


if __name__ == "__main__":
    main()
